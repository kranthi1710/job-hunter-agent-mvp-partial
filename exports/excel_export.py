from pathlib import Path
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from config import Config
from utils.date_utils import DateUtils


class ExcelExporter:
    """
    Responsible for exporting jobs to Excel.
    """

    def export(
        self,
        jobs
    ) -> str:

        Config.EXPORT_DIRECTORY.mkdir(
            exist_ok=True
        )

        workbook = Workbook()

        worksheet = workbook.active

        worksheet.title = "Jobs"

        headers = [

            "Company",

            "Job Title",

            "Location",

            "Employment Type",

            "Experience",

            "Salary",

            "Skills",

            "Posted Date",

            "Posted Age",

            "Application End Date",

            "Source",

            "Apply URL"

        ]

        worksheet.append(headers)

        self._style_header(
            worksheet
        )

        for job in jobs:

            worksheet.append([

                job.company,

                job.title,

                job.location,

                job.employment_type,

                job.experience_required,

                job.salary,

                ", ".join(job.skills),

                DateUtils.format_date(
                    job.posted_date
                ),

                DateUtils.age(
                    job.posted_date
                ),

                DateUtils.format_date(
                    job.application_end_date
                ),

                job.source,

                job.apply_url

            ])

        self._auto_size_columns(
            worksheet
        )

        worksheet.freeze_panes = "A2"

        worksheet.auto_filter.ref = (
            worksheet.dimensions
        )

        filename = (
            f"jobs_"
            f"{datetime.now():%Y%m%d_%H%M%S}.xlsx"
        )

        output_file = (
            Config.EXPORT_DIRECTORY
            /
            filename
        )

        workbook.save(
            output_file
        )

        return str(output_file)



    def _style_header(
        self,
        worksheet
    ):

        fill = PatternFill(

            fill_type="solid",

            fgColor="1F4E78"

        )

        font = Font(

            bold=True,

            color="FFFFFF"

        )

        for cell in worksheet[1]:

            cell.fill = fill

            cell.font = font

            cell.alignment = Alignment(
                horizontal="center"
            )



    def _auto_size_columns(
        self,
        worksheet
    ):

        for column in worksheet.columns:

            max_length = 0

            column_letter = get_column_letter(
                column[0].column
            )

            for cell in column:

                value = str(cell.value or "")

                if len(value) > max_length:

                    max_length = len(value)

            worksheet.column_dimensions[
                column_letter
            ].width = min(
                max_length + 3,
                50
            )